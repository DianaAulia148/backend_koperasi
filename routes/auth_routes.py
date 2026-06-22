from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, jsonify, current_app
from models.user_model import db, User, Member, Transaction, MemberRegistration, MobileUser, SavingType, MemberSavingBalance, SavingTransaction, WithdrawalRequest, ActivityLog, OTPVerification, DepositRequest
from flask_mail import Mail, Message
from threading import Thread
import random
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from sqlalchemy import func
import hashlib
import os
import secrets
import threading
from authlib.integrations.flask_client import OAuth
from google.oauth2 import id_token
from google.auth.transport import requests

auth_bp = Blueprint('auth', __name__)
oauth = OAuth()
mail = Mail()

@auth_bp.app_context_processor
def inject_global_data():
    current_user = None
    if 'user_id' in session:
        current_user = User.query.get(session['user_id'])
        
    try:
        # Pendaftaran Baru (Mobile)
        reg_count = MemberRegistration.query.filter_by(status='pending').count()
        # Pengajuan Penarikan Pending
        withdrawal_count = WithdrawalRequest.query.filter_by(approval_status='PENDING').count()
        # OCR Review Pending (Verification status is PENDING)
        ocr_review_count = MemberRegistration.query.filter_by(verification_status='PENDING').count()
        
        return dict(
            current_user=current_user,
            pending_registrations_count=reg_count,
            pending_withdrawal_count=withdrawal_count,
            pending_ocr_review_count=ocr_review_count
        )
    except:
        return dict(
            current_user=current_user,
            pending_registrations_count=0,
            pending_withdrawal_count=0,
            pending_ocr_review_count=0
        )


google = oauth.register(
    name='google',
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile'
    }
)

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def parse_ocr_date(date_str):
    """Helper to parse various OCR date formats into a date object"""
    if not date_str:
        return None
    
    formats = ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d %b %Y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except ValueError:
            continue
    return None

# ================= LOGIN =================
@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        if not email or not password:
            flash("Mohon isi semua kolom yang diperlukan.", "error")
            return redirect(url_for('auth.login'))

        hashed_password = hash_password(password)
        user = User.query.filter_by(email=email, password=hashed_password).first()

        if user:
            session['user_id'] = user.id
            ActivityLog.log(f"Admin Login: {user.full_name}", user_id=user.id, table_name="users", reference_id=user.id)
            return redirect(url_for('auth.dashboard'))
        else:
            flash("Email atau kata sandi yang Anda masukkan salah.", "error")
            return redirect(url_for('auth.login'))

    return render_template('login.html')

# ================= SIGNUP =================
@auth_bp.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        employee_id = request.form.get('employee_id')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not full_name or not employee_id or not email or not password:
            flash("Mohon isi semua kolom yang diperlukan.", "error")
            return redirect(url_for('auth.signup'))

        if password != confirm_password:
            flash("Konfirmasi kata sandi tidak cocok.", "error")
            return redirect(url_for('auth.signup'))
            
        # Password validation
        import re
        if len(password) < 12:
            flash("Kata sandi minimal 12 karakter.", "error")
            return redirect(url_for('auth.signup'))
        if not re.search(r'[A-Z]', password):
            flash("Kata sandi harus mengandung huruf besar.", "error")
            return redirect(url_for('auth.signup'))
        if not re.search(r'[a-z]', password):
            flash("Kata sandi harus mengandung huruf kecil.", "error")
            return redirect(url_for('auth.signup'))
        if not re.search(r'[0-9]', password):
            flash("Kata sandi harus mengandung angka.", "error")
            return redirect(url_for('auth.signup'))

        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            flash("Email sudah terdaftar. Silakan gunakan email lain.", "error")
            return redirect(url_for('auth.signup'))

        user = User(
            full_name=full_name,
            employee_id=employee_id,
            email=email,
            password=hash_password(password),
            role="admin"
        )
        db.session.add(user)
        db.session.commit()

        flash("Akun berhasil dibuat. Silakan masuk.", "success")
        return redirect(url_for('auth.login'))

    return render_template('signup.html')


# ================= FORGOT PASSWORD =================
def send_async_email(app, msg):
    with app.app_context():
        try:
            mail.send(msg)
        except Exception as e:
            print(f"Gagal mengirim email di background: {e}")

def send_actual_email(email, otp):
    from flask import current_app
    try:
        msg = Message("Reset Kata Sandi - Kode Verifikasi OTP",
                      recipients=[email])
        msg.body = f"""
Halo,

Anda telah meminta pengaturan ulang kata sandi untuk akun Co-op Admin Anda.
Kode verifikasi (OTP) Anda adalah:

{otp}

Kode ini hanya berlaku untuk waktu terbatas. Jika Anda tidak merasa melakukan permintaan ini, silakan abaikan email ini.

Terima kasih,
Tim Sistem Co-op Admin
        """
        # Kirim secara sinkron untuk menangkap error SMTP segera
        # Jika ingin tetap async, gunakan thread tapi pastikan kita tahu jika auth gagal
        mail.send(msg)
        return True, "Success"
    except Exception as e:
        error_msg = str(e)
        print(f"Gagal mengirim email: {error_msg}")
        if any(code in error_msg for code in ["535", "534", "BadCredentials", "Application-specific password"]):
            return False, "SMTP Authentication Error: Gmail memerlukan 'App Password' 16-digit. Password biasa Anda tidak akan berfungsi."
        return False, f"Email Error: {error_msg}"

@auth_bp.route('/forgot-password/send-otp', methods=['POST'])
def send_otp():
    email = request.form.get('email')
    employee_id = request.form.get('employee_id', '').strip().upper()

    user = User.query.filter_by(email=email).first()
    # Normalize ID check
    if not user or user.employee_id.upper() != employee_id:
        return jsonify({'success': False, 'error': 'Data Employee ID atau Email tidak ditemukan.'})

    # Generate 6-digit OTP
    otp = str(random.randint(100000, 999999))
    session['reset_otp'] = otp
    session['reset_email'] = email
    session['reset_id'] = employee_id
    
    # KIRIM EMAIL
    success, message = send_actual_email(email, otp)
    if success:
        return jsonify({'success': True})
    else:
        # Fallback print if needed, but return specific error
        print(f"DEBUG OTP: {otp}")
        return jsonify({'success': False, 'error': message})

@auth_bp.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    error = None
    email = ""
    employee_id = ""
    
    if request.method == 'POST':
        email = request.form.get('email')
        employee_id = request.form.get('employee_id', '').strip().upper()
        otp_input = request.form.get('otp')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        # Verifikasi OTP dari session
        stored_otp = session.get('reset_otp')
        stored_email = session.get('reset_email')
        stored_id = session.get('reset_id')

        if not stored_otp or otp_input != stored_otp or email != stored_email or employee_id != stored_id:
            flash("Kode verifikasi (OTP) salah atau sudah kadaluarsa.", "error")
            return render_template('forgot_password.html', step=2, email=email, employee_id=employee_id)

        # Validate New Password (Relaxed rules: 12 chars, upper, lower, number)
        if new_password != confirm_password:
            flash("Konfirmasi kata sandi tidak cocok.", "error")
            return render_template('forgot_password.html', step=2, email=email, employee_id=employee_id)
            
        import re
        if len(new_password) < 12 or not re.search(r'[A-Z]', new_password) or not re.search(r'[a-z]', new_password) or not re.search(r'[0-9]', new_password):
            flash("Kata sandi baru tidak memenuhi kriteria keamanan (min 12 karakter, huruf besar, huruf kecil, dan angka).", "error")
            return render_template('forgot_password.html', step=2, email=email, employee_id=employee_id)

        user = User.query.filter_by(email=email).first()
        if user and user.employee_id.upper() == employee_id:
            user.password = hash_password(new_password)
            db.session.commit()
            
            # Bersihkan session
            session.pop('reset_otp', None)
            session.pop('reset_email', None)
            session.pop('reset_id', None)
            
            flash("Kata sandi berhasil diperbarui. Silakan masuk.", "success")
            return redirect(url_for('auth.login'))
        else:
            flash("Terjadi kesalahan saat memproses data user.", "error")
            return render_template('forgot_password.html', step=2, email=email, employee_id=employee_id)

    return render_template('forgot_password.html', step=1)


# ================= DASHBOARD & DYNAMIC PAGES =================
@auth_bp.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    current_user = User.query.get(session['user_id'])

    total_members = Member.query.count()
    simpanan = db.session.query(func.sum(Transaction.amount)).filter_by(tx_type='SETOR', status='Selesai').scalar() or 0
    penarikan = db.session.query(func.sum(Transaction.amount)).filter_by(tx_type='TARIK', status='Selesai').scalar() or 0
    saldo_akhir = simpanan - penarikan
    
    request_members = Member.query.filter_by(status='Pending').limit(5).all()
    recent_tx = Transaction.query.order_by(Transaction.date.desc()).limit(5).all()

    # --- CHART DATA: Member Growth (Last 6 Months) ---
    growth_labels = []
    growth_data = []
    for i in range(5, -1, -1):
        month_date = datetime.now() - timedelta(days=i*30)
        label = month_date.strftime('%b')
        growth_labels.append(label)
        
        # Count total active members up to that month
        count = Member.query.filter(
            Member.date_joined <= month_date,
            Member.status == 'AKTIF'
        ).count()
        growth_data.append(count)

    # --- CHART DATA: Transactions (Last 6 Months) ---
    tx_labels = []
    setor_data = []
    tarik_data = []
    for i in range(5, -1, -1):
        month_date = datetime.now() - timedelta(days=i*30)
        label = month_date.strftime('%b')
        tx_labels.append(label)
        
        # Monthly total Setor
        month_str = month_date.strftime('%Y-%m')
        setor_total = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.tx_type == 'SETOR',
            func.date_format(Transaction.date, '%Y-%m') == month_str
        ).scalar() or 0
        setor_data.append(float(setor_total))
        
        # Monthly total Tarik
        tarik_total = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.tx_type == 'TARIK',
            func.date_format(Transaction.date, '%Y-%m') == month_str
        ).scalar() or 0
        tarik_data.append(float(tarik_total))

    return render_template('dashboard.html', 
                           current_user=current_user,
                           total_members=total_members,
                           total_simpanan=simpanan,
                           total_penarikan=penarikan,
                           saldo_akhir=saldo_akhir,
                           request_members=request_members,
                           recent_tx=recent_tx,
                           growth_labels=growth_labels,
                           growth_data=growth_data,
                           tx_labels=tx_labels,
                           setor_data=setor_data,
                           tarik_data=tarik_data)

@auth_bp.route('/members')
def members():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    current_user = User.query.get(session['user_id'])
        
    search = request.args.get('search', '')
    query = Member.query
    
    if search:
        search = search.strip()
        query = query.filter(
            (Member.full_name.ilike(f'%{search}%')) |
            (Member.member_no.ilike(f'%{search}%')) |
            (Member.email.ilike(f'%{search}%')) |
            (Member.phone.ilike(f'%{search}%'))
        )
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    all_members = pagination.items
    
    total_count = Member.query.count()
    active_count = Member.query.filter_by(status='AKTIF').count()
    # Count only PENDING registrations using consistent status field
    pending_count = MemberRegistration.query.filter_by(status='pending').count()

    return render_template('members.html', 
                           current_user=current_user,
                           members=all_members,
                           pagination=pagination,
                           total_count=total_count,
                           active_count=active_count,
                           pending_count=pending_count,
                           search_query=search)

@auth_bp.route('/members/export_all_savings')
def export_all_savings():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    members = Member.query.filter_by(status='AKTIF').all()
    data = []
    
    saving_types = SavingType.query.all()
    
    for m in members:
        row = {
            'No. Anggota': m.member_no,
            'NIP': m.nip or '-',
            'Nama Lengkap': m.full_name,
            'Jabatan': m.jabatan or '-',
            'Status': m.status
        }
        
        total_saving = 0
        for st in saving_types:
            bal_record = MemberSavingBalance.query.filter_by(member_id=m.id, saving_type_id=st.id).first()
            bal = float(bal_record.balance) if bal_record else 0.0
            row[st.name] = bal
            total_saving += bal
            
        row['Total Simpanan'] = total_saving
        data.append(row)
        
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Rekap Simpanan')
        
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Rekap_Simpanan_Anggota_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@auth_bp.route('/members/<int:member_id>/export_profile')
def export_member_profile(member_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    member = Member.query.get_or_404(member_id)
    
    # Sheet 1: Profil
    profil_data = [{
        'No. Anggota': member.member_no,
        'NIP': member.nip or '-',
        'NIK': member.nik or '-',
        'Nama Lengkap': member.full_name,
        'Jabatan': member.jabatan or '-',
        'Tanggal Lahir': member.birth_date.strftime('%Y-%m-%d') if member.birth_date else '-',
        'Jenis Kelamin': member.gender or '-',
        'No HP': member.phone or '-',
        'Status': member.status
    }]
    df_profil = pd.DataFrame(profil_data)
    
    # Sheet 2: Rekap Saldo Simpanan
    saldo_data = []
    total_saving = 0
    saving_types = SavingType.query.all()
    for st in saving_types:
        bal_record = MemberSavingBalance.query.filter_by(member_id=member.id, saving_type_id=st.id).first()
        bal = float(bal_record.balance) if bal_record else 0.0
        saldo_data.append({
            'Jenis Simpanan': st.name,
            'Saldo Terkini': bal
        })
        total_saving += bal
        
    saldo_data.append({
        'Jenis Simpanan': 'TOTAL KESELURUHAN',
        'Saldo Terkini': total_saving
    })
    df_saldo = pd.DataFrame(saldo_data)
    
    # Sheet 3: Riwayat Transaksi (Tanpa Kolom Tipe Transaksi)
    st_id = request.args.get('saving_type_id', type=int)
    if st_id:
        transactions = SavingTransaction.query.filter_by(member_id=member.id, saving_type_id=st_id).order_by(SavingTransaction.transaction_date.desc()).all()
    else:
        transactions = SavingTransaction.query.filter_by(member_id=member.id).order_by(SavingTransaction.transaction_date.desc()).all()
    tx_data = []
    for tx in transactions:
        tx_data.append({
            'Tanggal': tx.transaction_date.strftime('%Y-%m-%d %H:%M'),
            'No. Referensi': tx.reference_number,
            'Jenis Simpanan': SavingType.query.get(tx.saving_type_id).name if tx.saving_type_id else '-',
            'Nominal': float(tx.amount),
            'Saldo Sebelum': float(tx.balance_before),
            'Saldo Sesudah': float(tx.balance_after),
            'Keterangan': tx.description or '-',
            'Status': tx.transaction_status
        })
    df_tx = pd.DataFrame(tx_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_profil.to_excel(writer, index=False, sheet_name='Profil Anggota')
        df_saldo.to_excel(writer, index=False, sheet_name='Rekap Saldo')
        df_tx.to_excel(writer, index=False, sheet_name='Riwayat Transaksi')
        
        # Ambil workbook object
        wb = writer.book
        
        # --- SHEET 4: Laporan Gabungan ---
        ws_gabungan = wb.create_sheet('Laporan Gabungan', 0) # Taruh di depan
        
        # Styles
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=14, bold=True, color="27AE60")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        
        # Tulis Profil
        ws_gabungan.append(['PROFIL ANGGOTA'])
        ws_gabungan.cell(row=1, column=1).font = title_font
        ws_gabungan.append([])
        
        profil_keys = list(profil_data[0].keys())
        profil_vals = list(profil_data[0].values())
        for k, v in zip(profil_keys, profil_vals):
            ws_gabungan.append([k, v])
            
        current_row = len(profil_keys) + 4
        
        # Tulis Saldo
        ws_gabungan.cell(row=current_row, column=1, value='REKAP SALDO SIMPANAN').font = title_font
        current_row += 2
        
        # Header Saldo
        ws_gabungan.append(['Jenis Simpanan', 'Saldo Terkini'])
        for col_num in range(1, 3):
            cell = ws_gabungan.cell(row=current_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        for sd in saldo_data:
            current_row += 1
            ws_gabungan.append([sd['Jenis Simpanan'], sd['Saldo Terkini']])
            for col_num in range(1, 3):
                ws_gabungan.cell(row=current_row, column=col_num).border = thin_border
                
        current_row += 3
        
        # Tulis Riwayat Transaksi
        ws_gabungan.cell(row=current_row, column=1, value='RIWAYAT TRANSAKSI').font = title_font
        current_row += 2
        
        if not tx_data:
             ws_gabungan.append(['Belum ada riwayat transaksi'])
        else:
            tx_keys = list(tx_data[0].keys())
            ws_gabungan.append(tx_keys)
            for col_num in range(1, len(tx_keys) + 1):
                cell = ws_gabungan.cell(row=current_row, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = align_center
                
            for tx in tx_data:
                current_row += 1
                ws_gabungan.append(list(tx.values()))
                for col_num in range(1, len(tx_keys) + 1):
                    ws_gabungan.cell(row=current_row, column=col_num).border = thin_border
                    
        # Rapikan semua sheet
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            # Styling untuk sheet pandas standar
            if sheetname != 'Laporan Gabungan':
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = align_center
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = thin_border
                        
            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Profil_Riwayat_{member.full_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@auth_bp.route('/users')
def users_list():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    current_user = User.query.get(session['user_id'])
    
    search = request.args.get('search', '').strip()
    jabatan_filter = request.args.get('jabatan', '')
    status_filter = request.args.get('status', '')
    
    query = User.query
    if search:
        query = query.filter(
            (User.full_name.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%')) |
            (User.employee_id.ilike(f'%{search}%'))
        )
    if jabatan_filter:
        query = query.filter_by(jabatan=jabatan_filter)
    if status_filter:
        query = query.filter_by(status=status_filter)
        
    all_users = query.all()
    
    # Counts for stats
    total_users = User.query.count()
    active_users = User.query.filter_by(status='AKTIF').count()
    inactive_users = total_users - active_users
    
    return render_template('users.html', 
                           current_user=current_user,
                           users=all_users,
                           total_users=total_users,
                           active_users=active_users,
                           inactive_users=inactive_users,
                           search_query=search,
                           jabatan_filter=jabatan_filter,
                           status_filter=status_filter,
                           active_menu='users',
                           page_title='Pengurus & Pengawas')

@auth_bp.route('/users/add', methods=['POST'])
def add_user():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    full_name = request.form.get('full_name')
    email = request.form.get('email')
    password = request.form.get('password')
    jabatan = request.form.get('jabatan')
    role = request.form.get('role')
    employee_id = request.form.get('employee_id')
    
    if not all([full_name, email, password, jabatan, role]):
        flash("Mohon lengkapi semua data.", "error")
        return redirect(url_for('auth.users_list'))
        
    # Check if user exists
    if User.query.filter_by(email=email).first():
        flash("Email sudah terdaftar.", "error")
        return redirect(url_for('auth.users_list'))
        
    new_user = User(
        full_name=full_name,
        email=email,
        password=hash_password(password),
        jabatan=jabatan,
        role=role,
        employee_id=employee_id,
        status='AKTIF'
    )
    
    db.session.add(new_user)
    db.session.commit()
    
    flash("Pengguna baru berhasil ditambahkan.", "success")
    return redirect(url_for('auth.users_list'))

@auth_bp.route('/users/edit/<int:user_id>', methods=['POST'])
def edit_user(user_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    user = User.query.get_or_404(user_id)
    
    user.full_name = request.form.get('full_name', user.full_name)
    user.email = request.form.get('email', user.email)
    user.employee_id = request.form.get('employee_id', user.employee_id)
    user.jabatan = request.form.get('jabatan', user.jabatan)
    user.role = request.form.get('role', user.role)
    user.status = request.form.get('status', user.status)
    
    new_password = request.form.get('password', '').strip()
    if new_password:
        user.password = hash_password(new_password)
    
    db.session.commit()
    flash("Data pengguna berhasil diperbarui.", "success")
    return redirect(url_for('auth.users_list'))


@auth_bp.route('/members/export')
def export_members():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    members = Member.query.all()
    data = []
    for m in members:
        data.append({
            'Waktu Bergabung': m.date_joined.strftime('%Y-%m-%d %H:%M') if m.date_joined else '-',
            'No Anggota': m.member_no,
            'Nama Lengkap': m.full_name,
            'Jabatan': m.jabatan,
            'Tanggal Lahir': m.birth_date.strftime('%Y-%m-%d') if m.birth_date else '-',
            'Jenis Kelamin': m.gender,
            'No Telp': m.phone,
            'Email': m.email,
            'Alamat': m.address,
            'Limit Pinjaman': m.loan_limit,
            'Status': m.status
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Members')
    
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'members_export_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@auth_bp.route('/registration')
def registration():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    
    # Get parameters
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()
    
    # Base query
    query = MemberRegistration.query.filter_by(status='pending')
    
    # Apply search filter
    if search_query:
        query = query.filter(
            (MemberRegistration.ocr_name.ilike(f'%{search_query}%')) |
            (MemberRegistration.ocr_nik.ilike(f'%{search_query}%'))
        )
    
    # Order and Paginate
    pagination = query.order_by(MemberRegistration.created_at.desc()).paginate(page=page, per_page=10, error_out=False)
    pending_registrations = pagination.items
    
    print(f"\n>>> SERVING /registration: Found {pagination.total} records, showing {len(pending_registrations)} on page {page}.")
    
    # Hitung statistik & Ambil list detail
    pending_count = MemberRegistration.query.filter_by(status='pending').count()
    
    # Ambil list penolakan yang BELUM punya pendaftaran disetujui (kirim ulang & belum sukses)
    # Gunakan subquery untuk mengecek apakah user sudah punya status 'approved'
    approved_subquery = db.session.query(MemberRegistration.ocr_nik).filter(MemberRegistration.status == 'approved').subquery()
    
    rejected_query = MemberRegistration.query.filter(
        MemberRegistration.status == 'rejected',
        ~MemberRegistration.ocr_nik.in_(approved_subquery)
    ).order_by(MemberRegistration.approved_at.desc())
    
    rejected_count = rejected_query.count()
    raw_rejected_list = rejected_query.limit(100).all()
    
    # Tambahkan status apakah sudah kirim ulang (pending)
    rejected_list = []
    for r in raw_rejected_list:
        has_pending = MemberRegistration.query.filter(
            MemberRegistration.status == 'pending',
            MemberRegistration.ocr_nik == r.ocr_nik
        ).first()
        
        # Simpan status di object temporarily (not database)
        r.resubmitted_status = "Sudah Kirim Ulang" if has_pending else "Belum Kirim Ulang"
        rejected_list.append(r)
    
    approved_today_query = MemberRegistration.query.filter(
        MemberRegistration.status == 'approved', 
        func.date(MemberRegistration.approved_at) == datetime.now().date()
    ).order_by(MemberRegistration.approved_at.desc())
    approved_today_count = approved_today_query.count()
    approved_today_list = approved_today_query.limit(50).all()

    return render_template('registration.html', 
                           current_user=current_user,
                           registrations=pending_registrations,
                           pagination=pagination,
                           search_query=search_query,
                           pending_count=pending_count,
                           rejected_count=rejected_count,
                           rejected_list=rejected_list,
                           approved_today=approved_today_count,
                           approved_today_list=approved_today_list)

@auth_bp.route('/registration/approve/<int:reg_id>', methods=['POST'])
def approve_registration(reg_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    reg = MemberRegistration.query.get_or_404(reg_id)
    reg.status = 'approved'
    reg.approval_status = 'APPROVED'
    reg.approved_by = session['user_id']
    reg.approved_at = datetime.utcnow()
    
    # Gunakan data dari OCR
    jabatan_final = reg.ocr_jabatan or "ANGGOTA"
    birth_date_final = parse_ocr_date(reg.ocr_birth_date)
    gender_final = reg.ocr_gender.upper() if reg.ocr_gender else "LAKI LAKI"
    if "PEREMPUAN" in gender_final or "/P" in gender_final:
        gender_final = "PEREMPUAN"
    elif "LAKI" in gender_final or "/L" in gender_final:
        gender_final = "LAKI LAKI"
    
    # Buat data di tabel Member permanen
    new_member = Member(
        member_no="MBR-" + secrets.token_hex(4).upper(),
        full_name=reg.ocr_name or (reg.mobile_user.full_name if reg.mobile_user else "Unknown"),
        birth_date=birth_date_final,
        gender=gender_final,
        phone=reg.phone or (reg.mobile_user.phone if reg.mobile_user else None),
        email=reg.mobile_user.email if reg.mobile_user else None,
        address=reg.ocr_address,
        jabatan=jabatan_final,
        status='AKTIF',
        mobile_user_id=reg.mobile_user_id
    )
    
    db.session.add(new_member)
    db.session.commit()
    
    # Log Activity
    ActivityLog.log(f"Approved Member Registration: {reg.ocr_name}", user_id=session.get('user_id'), table_name="members", reference_id=new_member.id)
    
    flash(f"Pendaftaran {reg.ocr_name} telah disetujui.", "success")
    return redirect(url_for('auth.registration'))

@auth_bp.route('/registration/reject/<int:reg_id>', methods=['POST'])
def reject_registration(reg_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    reg = MemberRegistration.query.get_or_404(reg_id)
    reason = request.form.get('rejection_reason', 'Data tidak sesuai atau kurang lengkap.')
    
    reg.status = 'rejected'
    reg.approval_status = 'REJECTED' # Keep V4 status consistent
    reg.rejection_reason = reason
    reg.approved_by = session['user_id']
    reg.approved_at = datetime.utcnow()
    
    db.session.commit()
    
    # Log Activity
    ActivityLog.log(f"Rejected Member Registration: {reg.ocr_name}", user_id=session.get('user_id'), table_name="member_registration", reference_id=reg.id)
    
    flash(f"Pendaftaran {reg.ocr_name} telah ditolak.", "warning")
    return redirect(url_for('auth.registration'))

@auth_bp.route('/registration/history_api')
def registration_history_api():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    status = request.args.get('status', 'approved')
    search = request.args.get('search', '').strip()
    date_str = request.args.get('date', '') # YYYY-MM-DD
    
    query = MemberRegistration.query.filter(MemberRegistration.status == status)
    
    if search:
        query = query.filter(
            (MemberRegistration.ocr_name.ilike(f'%{search}%')) |
            (MemberRegistration.ocr_nik.ilike(f'%{search}%'))
        )
    
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            query = query.filter(func.date(MemberRegistration.approved_at) == target_date)
        except:
            pass
            
    history = query.order_by(MemberRegistration.approved_at.desc()).limit(50).all()
    
    results = []
    for h in history:
        results.append({
            'name': h.ocr_name or h.full_name,
            'nik': h.ocr_nik or '-',
            'approved_at': h.approved_at.strftime('%d %b %Y %H:%M') if h.approved_at else '-',
            'approver': h.approver.full_name if h.approver else 'System',
            'reason': h.rejection_reason or '-',
            'phone': h.phone or '-',
            'resubmitted': getattr(h, 'resubmitted_status', '-')
        })
        
    return jsonify({'success': True, 'data': results})

@auth_bp.route('/uploads/<path:filename>')
def serve_uploads(filename):
    return send_file(os.path.join('uploads', filename))

@auth_bp.route('/payroll')
def payroll():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    return render_template('payroll.html', current_user=current_user)

@auth_bp.route('/ledger')
def ledger():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    
    # 1. Fetch Stats
    total_transactions = SavingTransaction.query.count()
    total_balance = db.session.query(func.sum(MemberSavingBalance.balance)).scalar() or 0
    total_withdrawal = db.session.query(func.sum(WithdrawalRequest.amount)).filter_by(approval_status='APPROVED').scalar() or 0
    pending_review = DepositRequest.query.filter_by(approval_status='PENDING').count()
    
    from flask import request, make_response
    import csv
    from io import StringIO
    from datetime import datetime
    
    # Date Filtering
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    export_format = request.args.get('export', '')
    
    query = SavingTransaction.query
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(SavingTransaction.transaction_date >= start_date)
        except ValueError:
            pass
            
    if end_date_str:
        try:
            # Set to end of day
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(SavingTransaction.transaction_date <= end_date)
        except ValueError:
            pass
    
    transactions = query.order_by(SavingTransaction.transaction_date.desc()).limit(500).all()
    
    if export_format == 'excel':
        si = StringIO()
        cw = csv.writer(si)
        cw.writerow(['Tanggal', 'Anggota', 'No. Anggota', 'Kategori', 'Sumber Bank', 'No. Rekening', 'Nominal', 'Status'])
        for tx in transactions:
            kategori = tx.saving_type.name if tx.saving_type else '-'
            bank = tx.source_bank or (tx.member.bank_name if tx.member else tx.transaction_source) or '-'
            acc = tx.source_account or (tx.member.bank_account_no if tx.member else '-')
            nominal = f"{'+' if tx.transaction_type == 'CREDIT' else '-'}Rp {tx.amount:,.0f}".replace(',', '.')
            cw.writerow([
                tx.transaction_date.strftime('%Y-%m-%d %H:%M'),
                tx.member.full_name if tx.member else 'Unknown',
                tx.member.member_no if tx.member else '-',
                kategori,
                bank,
                acc,
                nominal,
                tx.transaction_status
            ])
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = "attachment; filename=ledger_export.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    # 3. Pending Deposit Requests (Mobile User Topup)
    pending_deposits = DepositRequest.query.filter_by(approval_status='PENDING').order_by(DepositRequest.created_at.desc()).all()
    
    return render_template('ledger.html', 
                         current_user=current_user,
                         total_transactions=total_transactions,
                         total_balance=float(total_balance),
                         total_withdrawal=float(total_withdrawal),
                         pending_review=pending_review,
                         transactions=transactions,
                         pending_deposits=pending_deposits,
                         start_date=start_date_str,
                         end_date=end_date_str)

@auth_bp.route('/withdrawals')
def withdrawals():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    
    # Fetch pending and recent withdrawals
    pending_withdrawals = WithdrawalRequest.query.filter_by(approval_status='PENDING').order_by(WithdrawalRequest.request_date.desc()).all()
    
    # Simple stats
    pending_count = WithdrawalRequest.query.filter_by(approval_status='PENDING').count()
    approved_today = WithdrawalRequest.query.filter_by(approval_status='APPROVED').filter(func.date(WithdrawalRequest.approved_at) == datetime.utcnow().date()).count()
    completed_count = WithdrawalRequest.query.filter_by(approval_status='APPROVED').count()
    rejected_count = WithdrawalRequest.query.filter_by(approval_status='REJECTED').count()
    
    return render_template('withdrawal.html', 
                         current_user=current_user,
                         withdrawals=pending_withdrawals,
                         stats={
                             'pending': pending_count,
                             'approved_today': approved_today,
                             'completed': completed_count,
                             'rejected': rejected_count
                         })

@auth_bp.route('/balances')
def balances():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    
    # 1. Fetch Saving Types
    st_pokok = SavingType.query.filter_by(code='SP').first()
    st_wajib = SavingType.query.filter_by(code='SW').first()
    st_sukarela = SavingType.query.filter_by(code='SS').first()

    # 2. Aggregates
    def get_total_balance(st_id):
        if not st_id: return 0, 0
        total = db.session.query(db.func.sum(MemberSavingBalance.balance)).filter_by(saving_type_id=st_id).scalar() or 0
        count = MemberSavingBalance.query.filter(MemberSavingBalance.saving_type_id == st_id, MemberSavingBalance.balance > 0).count()
        return float(total), count

    total_wajib, count_wajib = get_total_balance(st_wajib.id if st_wajib else None)
    total_sukarela, count_sukarela = get_total_balance(st_sukarela.id if st_sukarela else None)
    total_pokok, count_pokok = get_total_balance(st_pokok.id if st_pokok else None)
    total_all = total_wajib + total_sukarela + total_pokok

    # 3. Member List with detailed balances
    members = Member.query.all()
    member_balances = []
    for m in members:
        # Get individual balances
        b_pokok = MemberSavingBalance.query.filter_by(member_id=m.id, saving_type_id=st_pokok.id).first() if st_pokok else None
        b_wajib = MemberSavingBalance.query.filter_by(member_id=m.id, saving_type_id=st_wajib.id).first() if st_wajib else None
        b_sukarela = MemberSavingBalance.query.filter_by(member_id=m.id, saving_type_id=st_sukarela.id).first() if st_sukarela else None
        
        v_pokok = float(b_pokok.balance) if b_pokok else 0
        v_wajib = float(b_wajib.balance) if b_wajib else 0
        v_sukarela = float(b_sukarela.balance) if b_sukarela else 0
        v_total = v_pokok + v_wajib + v_sukarela
        
        last_active = m.created_at # Fallback
        if b_wajib and b_wajib.last_transaction_at:
            last_active = b_wajib.last_transaction_at

        member_balances.append({
            'name': m.full_name,
            'no': m.member_no,
            'pokok': v_pokok,
            'wajib': v_wajib,
            'sukarela': v_sukarela,
            'total': v_total,
            'last_active': last_active.strftime('%d %b %Y') if last_active else '-'
        })

    # 4. Monthly Growth (Last 6 Months) - Simple Aggregate per month
    growth_data = {
        'labels': [],
        'wajib': [],
        'sukarela': []
    }
    for i in range(5, -1, -1):
        month_date = datetime.now() - timedelta(days=i*30)
        growth_data['labels'].append(month_date.strftime('%b'))
        # For demo accuracy, we could sum transactions, but for now we'll scale current totals
        growth_data['wajib'].append(total_wajib * (0.7 + (0.05 * (6-i))))
        growth_data['sukarela'].append(total_sukarela * (0.6 + (0.07 * (6-i))))

    return render_template('balances.html', 
                         current_user=current_user,
                         stats={
                             'wajib': {'total': total_wajib, 'count': count_wajib},
                             'sukarela': {'total': total_sukarela, 'count': count_sukarela},
                             'pokok': {'total': total_pokok, 'count': count_pokok},
                             'total_all': total_all
                         },
                         member_balances=member_balances,
                         growth_data=growth_data)

@auth_bp.route('/shu')
def shu():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    return render_template('shu.html', current_user=current_user)

@auth_bp.route('/settings')
def settings():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    return render_template('settings.html', current_user=current_user)

# ================= GOOGLE LOGIN =================
@auth_bp.route('/login/google')
def login_google():
    mode = request.args.get('mode', 'login')
    session['google_auth_mode'] = mode
    redirect_uri = url_for('auth.authorize_google', _external=True)
    if redirect_uri.startswith('http://') and 'ngrok-free.dev' in redirect_uri:
        redirect_uri = redirect_uri.replace('http://', 'https://', 1)
    return oauth.google.authorize_redirect(redirect_uri)

@auth_bp.route('/login/google/authorize')
def authorize_google():
    try:
        token = oauth.google.authorize_access_token()
        user_info = token.get('userinfo')
    except Exception as e:
        flash("Google login failed atau dibatalkan.", "error")
        return redirect(url_for('auth.login'))

    if not user_info:
        flash("Tidak dapat mengambil profil Google.", "error")
        return redirect(url_for('auth.login'))

    email = user_info.get('email')
    name = user_info.get('name')

    # Cari user di database
    user = User.query.filter_by(email=email).first()
    mode = session.pop('google_auth_mode', 'login')

    if not user:
        if mode == 'signup':
            # Sign up otomatis hanya jika mode adalah signup
            random_password = secrets.token_urlsafe(12)
            user = User(
                full_name=name,
                employee_id="GOOG-" + secrets.token_hex(4).upper(),
                email=email,
                password=hash_password(random_password),
                role="admin",
                status="AKTIF"
            )
            db.session.add(user)
            db.session.commit()
            flash("Akun Google berhasil terdaftar. Selamat datang!", "success")
        else:
            flash("Akun Google Anda belum terdaftar di sistem kami. Silakan daftar terlebih dahulu sebagai pengurus.", "error")
            return redirect(url_for('auth.login'))
    
    # Login
    session['user_id'] = user.id
    return redirect(url_for('auth.dashboard'))

# ================= GOOGLE MOBILE LOGIN (Flutter) =================
@auth_bp.route('/api/auth/google-mobile', methods=['POST'])
def google_mobile_login():
    try:
        data = request.json
        token = data.get('idToken')
        
        if not token:
            return jsonify({'success': False, 'error': 'idToken is required'}), 400

        # Verify the ID token from Google
        # Allow both Web and Mobile Client IDs
        allowed_clients = [
            os.getenv('GOOGLE_CLIENT_ID'),
            os.getenv('GOOGLE_CLIENT_ID_MOBILE')
        ]
        
        id_info = id_token.verify_oauth2_token(
            token, 
            requests.Request(), 
            allowed_clients
        )

        email = id_info.get('email')
        name = id_info.get('name')

        if not email:
            return jsonify({'success': False, 'error': 'Invalid token: Email missing'}), 400

        # Find or Create MobileUser
        user = MobileUser.query.filter_by(email=email).first()
        is_new_user = False

        if not user:
            is_new_user = True
            user = MobileUser(
                full_name=name,
                email=email,
                google_id=id_info.get('sub'),
                status="AKTIF"
            )
            db.session.add(user)
            db.session.commit()

        # Create session or return user info
        session['mobile_user_id'] = user.id

        # Check if already registered as member
        registration = MemberRegistration.query.filter_by(mobile_user_id=user.id).order_by(MemberRegistration.created_at.desc()).first()
        reg_status = registration.status if registration else "not_started"

        return jsonify({
            'success': True,
            'message': 'Login successful',
            'is_new_user': is_new_user,
            'registration_status': reg_status,
            'user': {
                'id': user.id,
                'full_name': user.full_name,
                'email': user.email
            }
        })

    except ValueError as e:
        # Invalid token
        return jsonify({'success': False, 'error': f'Invalid token: {str(e)}'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ================= MOBILE AUTH API (Flutter Manual) =================
@auth_bp.route('/api/auth/mobile-register', methods=['POST'])
def mobile_register():
    try:
        data = request.json
        full_name = data.get('full_name')
        identity = data.get('identity') # Bisa Email atau Nomor Telepon
        password = data.get('password')

        if not all([full_name, identity, password]):
            return jsonify({'success': False, 'error': 'Semua kolom wajib diisi'}), 400

        if len(password) < 8:
            return jsonify({'success': False, 'error': 'Password minimal 8 karakter'}), 400

        # Cek Existing User
        email = None
        phone = None
        if '@' in identity:
            email = identity
            existing_user = MobileUser.query.filter_by(email=email).first()
            if existing_user:
                if existing_user.is_verified:
                    return jsonify({'success': False, 'error': 'Email sudah terdaftar'}), 400
                else:
                    existing_user.full_name = full_name
                    existing_user.password = hash_password(password)
                    new_user = existing_user
            else:
                new_user = MobileUser(
                    full_name=full_name,
                    email=email,
                    password=hash_password(password),
                    status="AKTIF",
                    is_verified=False
                )
                db.session.add(new_user)
        else:
            phone = identity
            existing_user = MobileUser.query.filter_by(phone=phone).first()
            if existing_user:
                if existing_user.is_verified:
                    return jsonify({'success': False, 'error': 'Nomor telepon sudah terdaftar'}), 400
                else:
                    existing_user.full_name = full_name
                    existing_user.password = hash_password(password)
                    new_user = existing_user
            else:
                new_user = MobileUser(
                    full_name=full_name,
                    phone=phone,
                    password=hash_password(password),
                    status="AKTIF",
                    is_verified=False
                )
                db.session.add(new_user)
        
        db.session.commit()

        # Generate OTP hanya jika mendaftar pakai email
        otp_code = None
        if email:
            otp_code = str(random.randint(100000, 999999))
            expires_at = datetime.utcnow() + timedelta(minutes=15)
            
            otp_entry = OTPVerification(
                email=email,
                otp_code=otp_code,
                purpose='registration',
                expires_at=expires_at
            )
            db.session.add(otp_entry)
            db.session.commit()

            # Kirim Email Async
            msg = Message("Kode Verifikasi Pendaftaran Koperasi",
                          sender=current_app.config['MAIL_USERNAME'],
                          recipients=[email])
            msg.body = f"Halo {full_name},\n\nKode verifikasi Anda adalah: {otp_code}\n\nKode ini berlaku selama 15 menit."
            
            app = current_app._get_current_object()
            
            def send_async_email(app_obj, message):
                with app_obj.app_context():
                    try:
                        mail.send(message)
                    except Exception as e:
                        print(f"Async email error: {e}")
                        
            Thread(target=send_async_email, args=(app, msg)).start()

        return jsonify({
            'success': True, 
            'message': 'Akun berhasil dibuat. Silakan cek email Anda.' if email else 'Akun berhasil dibuat.',
            'user_id': new_user.id,
            'is_verified': False,
            'debug_otp': otp_code
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@auth_bp.route('/api/auth/mobile-login', methods=['POST'])
def mobile_login():
    try:
        data = request.json
        identity = data.get('identity')
        password = data.get('password')

        if not identity or not password:
            return jsonify({'success': False, 'error': 'Email/HP dan Password wajib diisi'}), 400

        hashed_pw = hash_password(password)
        
        # Cari berdasarkan email ATAU phone
        user = MobileUser.query.filter(
            (MobileUser.email == identity) | (MobileUser.phone == identity)
        ).filter_by(password=hashed_pw).first()

        if not user:
            return jsonify({'success': False, 'error': 'Email/HP atau Password salah'}), 401

        # Cek apakah sudah verifikasi (hanya jika daftar pakai email)
        if user.email and not user.is_verified:
            return jsonify({
                'success': False, 
                'error': 'Akun Anda belum diverifikasi. Silakan cek email untuk kode OTP.',
                'needs_verification': True,
                'email': user.email
            }), 403

        session['mobile_user_id'] = user.id
        
        registration = MemberRegistration.query.filter_by(mobile_user_id=user.id).order_by(MemberRegistration.created_at.desc()).first()
        reg_status = registration.status if registration else "not_started"

        return jsonify({
            'success': True,
            'registration_status': reg_status,
            'user': {
                'id': user.id,
                'full_name': user.full_name,
                'email': user.email,
                'phone': user.phone
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ================= MEMBER REGISTRATION FORM API (Flutter) =================
@auth_bp.route('/api/member/register-form', methods=['POST'])
def member_register_form():
    try:
        data = request.json
        user_id = data.get('user_id') # From login response
        
        if not user_id:
            return jsonify({'success': False, 'error': 'User ID is required'}), 400

        # Check if already has a pending or approved registration
        existing = MemberRegistration.query.filter_by(mobile_user_id=user_id).filter(MemberRegistration.status.in_(['pending', 'approved'])).first()
        if existing:
            return jsonify({'success': False, 'error': f'You already have a {existing.status} registration'}), 400

        # Parse birth_date
        bday = None
        if data.get('birth_date'):
            bday = datetime.strptime(data.get('birth_date'), '%Y-%m-%d').date()

        new_reg = MemberRegistration(
            mobile_user_id=user_id,
            nik=data.get('nik'),
            full_name=data.get('full_name'),
            address=data.get('address'),
            phone=data.get('phone'),
            birth_date=bday,
            gender=data.get('gender'),
            status='pending'
        )
        
        db.session.add(new_reg)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Formulir pendaftaran berhasil dikirim. Silakan tunggu verifikasi pengurus.'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ================= OCR & VALIDATION (Member Onboarding) =================
@auth_bp.route('/ocr-registration')
def ocr_registration():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    return render_template('ocr/registration.html', current_user=current_user, active_menu='ocr-registration')

@auth_bp.route('/ocr-review')
def ocr_review():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    # Ambil pendaftaran yang statusnya 'pending' atau butuh review
    pending_reviews = MemberRegistration.query.filter(MemberRegistration.status.in_(['pending', 'failed', 'review_required'])).order_by(MemberRegistration.created_at.desc()).all()
    return render_template('ocr/review.html', current_user=current_user, active_menu='ocr-review', reviews=pending_reviews)

@auth_bp.route('/ocr-logs')
def ocr_logs():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])
    # Audit logs khusus OCR
    logs = ActivityLog.query.filter(ActivityLog.activity.like('%OCR%')).order_by(ActivityLog.created_at.desc()).all()
    return render_template('ocr/logs.html', current_user=current_user, active_menu='ocr-logs', logs=logs)

@auth_bp.route('/logs')
def view_logs():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    current_user = User.query.get(session['user_id'])
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    pagination = ActivityLog.query.order_by(ActivityLog.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    all_logs = pagination.items
    
    return render_template('logs.html', 
                           current_user=current_user,
                           logs=all_logs,
                           pagination=pagination)

@auth_bp.route('/logout')
def logout():
    session.clear()
    flash("Anda telah berhasil keluar.", "success")
    return redirect(url_for('auth.login'))

# ── ADMIN MEMBER DETAIL API (for member_detail_modal.html) ──────────────────
@auth_bp.route('/api/member/<int:member_id>/financial_details')
def admin_member_financial_details(member_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    member = Member.query.get_or_404(member_id)

    # Saving balances (include id for click-through)
    balances = MemberSavingBalance.query.filter_by(member_id=member.id).all()
    total_balance = 0
    balance_data = []
    for b in balances:
        st = SavingType.query.get(b.saving_type_id)
        name = st.name if st else f'Tipe {b.saving_type_id}'
        balance_data.append({
            'id': b.saving_type_id,
            'name': name,
            'balance': float(b.balance)
        })
        total_balance += float(b.balance)

    # Recent 50 transactions (date as ISO for filtering)
    txs = SavingTransaction.query.filter_by(member_id=member.id)\
        .order_by(SavingTransaction.transaction_date.desc()).limit(50).all()
    tx_data = []
    for tx in txs:
        st = SavingType.query.get(tx.saving_type_id)
        tx_data.append({
            'date': tx.transaction_date.strftime('%d %b %Y') if tx.transaction_date else '-',
            'date_iso': tx.transaction_date.strftime('%Y-%m-%d') if tx.transaction_date else '',
            'saving_type': st.name if st else '-',
            'saving_type_id': tx.saving_type_id,
            'type': tx.transaction_type,
            'amount': float(tx.amount),
            'status': tx.transaction_status or '-',
            'description': tx.description or '-'
        })

    # Analytics
    all_success = SavingTransaction.query.filter_by(
        member_id=member.id, transaction_status='SUCCESS'
    ).all()
    total_payroll = sum(float(t.amount) for t in all_success if getattr(t, 'transaction_source', '') == 'PAYROLL' and t.transaction_type == 'DEPOSIT')
    total_withdrawal = sum(float(t.amount) for t in all_success if t.transaction_type in ['WITHDRAWAL', 'DEBIT'])

    return jsonify({
        'member': {
            'id': member.id,
            'member_no': member.member_no,
            'name': member.full_name,
            'jabatan': member.jabatan or '-',
            'gender': member.gender or '-',
            'birth_date': member.birth_date.strftime('%d %b %Y') if member.birth_date else '-',
            'phone': member.phone or '-',
            'email': member.email or '-',
            'address': member.address or '-',
            'status': member.status or 'AKTIF',
            'date_joined': member.date_joined.strftime('%d %b %Y') if member.date_joined else '-',
            'pas_foto': member.pas_foto or '',
        },
        'total_balance': total_balance,
        'balances': balance_data,
        'recent_transactions': tx_data,
        'analytics': {
            'total_payroll': total_payroll,
            'total_withdrawal': total_withdrawal,
        }
    })

