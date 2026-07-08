from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, send_file, flash
from models.user_model import (db, User, MemberSavingBalance, SavingTransaction,
                                PayrollBatch, WithdrawalRequest, DepositRequest,
                                Member, SavingType, PayrollBatchDetail, Notification)
from datetime import datetime, timezone
from sqlalchemy import func
import secrets
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

finance_bp = Blueprint('finance', __name__)


# ─────────────────────────────────────────────────────────────────────────────
# PAYROLL HALAMAN UTAMA
# ─────────────────────────────────────────────────────────────────────────────
@finance_bp.route('/finance/payroll')
def payroll():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])

    batches = PayrollBatch.query.order_by(PayrollBatch.uploaded_at.desc()).all()

    total_payroll_this_month = db.session.query(func.sum(PayrollBatch.total_amount)).filter(
        func.extract('month', PayrollBatch.uploaded_at) == datetime.now().month,
        func.extract('year', PayrollBatch.uploaded_at) == datetime.now().year
    ).scalar() or 0

    total_batches = PayrollBatch.query.count()
    successful_distributions = PayrollBatchDetail.query.filter_by(distribution_status='SUCCESS').count()
    active_members_count = Member.query.filter_by(status='AKTIF').count()

    stats = {
        'total_payroll': total_payroll_this_month,
        'total_batches': total_batches,
        'successful_distributions': successful_distributions,
        'active_members': active_members_count,
    }

    return render_template('payroll.html',
                           current_user=current_user,
                           active_menu='payroll',
                           page_title='Payroll Automation',
                           batches=batches,
                           stats=stats)


# ─────────────────────────────────────────────────────────────────────────────
# API: PROSES PAYROLL OTOMATIS
# ─────────────────────────────────────────────────────────────────────────────
@finance_bp.route('/finance/payroll/auto_process', methods=['POST'])
def auto_process_payroll():
    """
    Proses payroll otomatis: distribusi simpanan wajib ke semua anggota aktif.
    Body JSON (opsional): { "amount_per_member": 500000 }
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    try:
        data = request.get_json(silent=True) or {}
        amount_per_member = int(data.get('amount_per_member', 500000))

        active_members = Member.query.filter_by(status='AKTIF').all()
        if not active_members:
            return jsonify({'success': False, 'error': 'Tidak ada anggota aktif.'}), 400

        st_wajib = SavingType.query.filter_by(code='SW').first()
        if not st_wajib:
            return jsonify({'success': False, 'error': 'Tipe simpanan Wajib (kode SW) tidak ditemukan.'}), 400

        # Cegah duplikat proses bulan yang sama
        existing = PayrollBatch.query.filter_by(
            period_month=datetime.now().month,
            period_year=datetime.now().year,
            distribution_status='SUCCESS'
        ).first()
        if existing:
            return jsonify({
                'success': False,
                'error': f'Payroll bulan ini sudah diproses (Batch: {existing.batch_code}).'
            }), 400

        total_amount = len(active_members) * amount_per_member
        batch_code = "PR-" + datetime.now().strftime("%Y%m") + "-" + secrets.token_hex(2).upper()

        new_batch = PayrollBatch(
            batch_code=batch_code,
            period_month=datetime.now().month,
            period_year=datetime.now().year,
            total_amount=total_amount,
            total_members=len(active_members),
            distribution_status='PROCESSING',
            validation_status='SUCCESS',
            uploaded_by=session['user_id'],
            uploaded_at=datetime.now(timezone.utc),
        )
        db.session.add(new_batch)
        db.session.flush()

        success_count = 0
        failed_count = 0

        for member in active_members:
            try:
                # Cek/buat saldo simpanan wajib
                balance_record = MemberSavingBalance.query.filter_by(
                    member_id=member.id, saving_type_id=st_wajib.id
                ).first()
                if balance_record and balance_record.status == 'INACTIVE':
                    continue
                if not balance_record:
                    balance_record = MemberSavingBalance(
                        member_id=member.id, saving_type_id=st_wajib.id, balance=0
                    )
                    db.session.add(balance_record)
                    db.session.flush()

                balance_before = float(balance_record.balance)
                balance_after = balance_before + amount_per_member

                # Update saldo
                balance_record.balance = balance_after
                balance_record.last_transaction_at = datetime.now(timezone.utc)

                # Buat detail payroll
                detail = PayrollBatchDetail(
                    payroll_batch_id=new_batch.id,
                    member_id=member.id,
                    saving_type_id=st_wajib.id,
                    amount=amount_per_member,
                    distribution_status='SUCCESS'
                )
                db.session.add(detail)
                db.session.flush()

                # Buat transaksi dengan semua kolom wajib
                trx = SavingTransaction(
                    member_id=member.id,
                    payroll_batch_detail_id=detail.id,
                    saving_type_id=st_wajib.id,
                    transaction_type='DEBIT',
                    amount=amount_per_member,
                    balance_before=balance_before,
                    balance_after=balance_after,
                    transaction_source='PAYROLL',
                    reference_number="TRX-PAY-" + secrets.token_hex(6).upper(),
                    transaction_date=datetime.now(timezone.utc),
                    transaction_status='SUCCESS',
                    processed_by=session['user_id'],
                    processed_at=datetime.now(timezone.utc),
                    description=f'Pencatatan Simpanan Wajib oleh pengurus koperasi'
                )
                db.session.add(trx)
                
                Notification.create(
                    member_id=member.id,
                    title="Simpanan Masuk (Payroll)",
                    message=f"Simpanan Wajib sebesar Rp {amount_per_member:,} telah berhasil diproses oleh pengurus koperasi dan ditambahkan ke saldo simpanan Anda.".replace(',', '.'),
                    notification_type="PAYROLL"
                )

                success_count += 1

            except Exception as member_err:
                failed_count += 1
                print(f"[PAYROLL] Gagal anggota {member.id}: {member_err}")

        # Update status final batch
        new_batch.distribution_status = 'SUCCESS' if failed_count == 0 else 'PARTIAL'
        new_batch.success_count = success_count
        new_batch.failed_count = failed_count
        new_batch.total_members = success_count
        new_batch.total_amount = success_count * amount_per_member
        new_batch.processed_at = datetime.now(timezone.utc)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Payroll berhasil: {success_count} anggota. Gagal: {failed_count}.',
            'batch_code': batch_code,
            'total_members': len(active_members),
            'success_count': success_count,
            'failed_count': failed_count,
            'total_amount': success_count * amount_per_member
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# API: STATISTIK REAL-TIME (untuk di-polling frontend)
# ─────────────────────────────────────────────────────────────────────────────
@finance_bp.route('/finance/payroll/stats')
def payroll_stats():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    total_payroll_this_month = db.session.query(func.sum(PayrollBatch.total_amount)).filter(
        func.extract('month', PayrollBatch.uploaded_at) == datetime.now().month,
        func.extract('year', PayrollBatch.uploaded_at) == datetime.now().year
    ).scalar() or 0

    total_batches = PayrollBatch.query.count()
    successful_distributions = PayrollBatchDetail.query.filter_by(distribution_status='SUCCESS').count()
    failed_distributions = PayrollBatchDetail.query.filter_by(distribution_status='FAILED').count()
    active_members = Member.query.filter_by(status='AKTIF').count()
    latest_batch = PayrollBatch.query.order_by(PayrollBatch.uploaded_at.desc()).first()

    return jsonify({
        'success': True,
        'total_payroll_this_month': float(total_payroll_this_month),
        'total_batches': total_batches,
        'successful_distributions': successful_distributions,
        'failed_distributions': failed_distributions,
        'active_members': active_members,
        'latest_batch_code': latest_batch.batch_code if latest_batch else None,
        'latest_batch_status': latest_batch.distribution_status if latest_batch else None,
        'latest_success_count': latest_batch.success_count if latest_batch else 0,
        'latest_failed_count': latest_batch.failed_count if latest_batch else 0,
        'latest_total_members': latest_batch.total_members if latest_batch else 0,
    })


# ─────────────────────────────────────────────────────────────────────────────
# API: DETAIL BATCH (untuk modal)
# ─────────────────────────────────────────────────────────────────────────────
@finance_bp.route('/finance/payroll/batch/<batch_code>/detail')
def batch_detail_api(batch_code):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    batch = PayrollBatch.query.filter_by(batch_code=batch_code).first()
    if not batch:
        return jsonify({'success': False, 'error': 'Batch tidak ditemukan'}), 404

    details = PayrollBatchDetail.query.filter_by(payroll_batch_id=batch.id).all()
    result = []
    for d in details:
        member = Member.query.get(d.member_id)
        result.append({
            'member_no': member.member_no if member else '-',
            'member_name': member.full_name if member else 'Unknown',
            'amount': float(d.amount),
            'status': d.distribution_status,
        })

    success_count = sum(1 for d in details if d.distribution_status == 'SUCCESS')
    failed_count = sum(1 for d in details if d.distribution_status == 'FAILED')

    return jsonify({
        'success': True,
        'batch_code': batch.batch_code,
        'period': f"Bulan {batch.period_month} Tahun {batch.period_year}",
        'total_members': batch.total_members,
        'total_amount': float(batch.total_amount),
        'success_count': success_count,
        'failed_count': failed_count,
        'distribution_status': batch.distribution_status,
        'uploaded_at': batch.uploaded_at.strftime('%d %b %Y %H:%M') if batch.uploaded_at else '-',
        'details': result
    })


# ─────────────────────────────────────────────────────────────────────────────
# SIMPANAN (LEDGER)
# ─────────────────────────────────────────────────────────────────────────────
@finance_bp.route('/finance/savings')
def savings():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])

    from datetime import date
    today = date.today()
    first_day_of_month = today.replace(day=1)

    # Total semua transaksi (sepanjang waktu)
    total_transactions = SavingTransaction.query.filter(
        SavingTransaction.deleted_at.is_(None)
    ).count()
    # Total transaksi bulan ini
    total_transactions_bulan_ini = SavingTransaction.query.filter(
        SavingTransaction.deleted_at.is_(None),
        SavingTransaction.transaction_date >= first_day_of_month
    ).count()

    # Saldo simpanan saat ini (snapshot terkini dari tabel balance)
    total_balance = db.session.query(func.sum(MemberSavingBalance.balance)).scalar() or 0

    # Total penerimaan (setoran) sepanjang waktu — beda dengan saldo
    total_deposit_alltime = db.session.query(func.sum(SavingTransaction.amount)).filter(
        SavingTransaction.transaction_type == 'DEBIT',
        SavingTransaction.transaction_status == 'SUCCESS',
        SavingTransaction.deleted_at.is_(None)
    ).scalar() or 0

    # Total penarikan sepanjang waktu — hanya penarikan sukses pada transaksi simpanan
    total_withdrawal = db.session.query(func.sum(SavingTransaction.amount)).filter(
        SavingTransaction.transaction_type == 'CREDIT',
        SavingTransaction.transaction_source == 'WITHDRAWAL',
        SavingTransaction.transaction_status == 'SUCCESS',
        SavingTransaction.deleted_at.is_(None)
    ).scalar() or 0

    pending_review = DepositRequest.query.filter_by(approval_status='PENDING').count()

    search_query = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Tampilkan semua anggota aktif (tanpa pengecualian berdasarkan nama)
    query = Member.query.filter(Member.deleted_at.is_(None))
    
    if search_query:
        query = query.filter(db.or_(
            Member.full_name.ilike(f'%{search_query}%'),
            Member.member_no.ilike(f'%{search_query}%')
        ))
        
    pagination = query.order_by(Member.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    members = pagination.items
    
    saving_types = SavingType.query.all()
    member_balances = {}
    member_applied = {}
    member_banks = {}
    member_accounts = {}
    member_totals = {}
    member_statuses = {}
    for m in members:
        # Find approved/pending deposit requests to get bank & account number
        deps = DepositRequest.query.filter(
            DepositRequest.member_id == m.id,
            DepositRequest.approval_status.in_(['APPROVED', 'PENDING'])
        ).order_by(DepositRequest.created_at.desc()).all()
        
        db_balances = MemberSavingBalance.query.filter_by(member_id=m.id).all()
        balances_map = {b.saving_type_id: float(b.balance) for b in db_balances}
        statuses_map = {b.saving_type_id: getattr(b, 'status', 'ACTIVE') or 'ACTIVE' for b in db_balances}
        member_balances[m.id] = balances_map
        member_totals[m.id] = sum(balances_map.values())
        member_statuses[m.id] = statuses_map
        
        applied = {}
        banks = {}
        accounts = {}
        for st in saving_types:
            st_deps = [d for d in deps if d.saving_type_id == st.id]
            if st_deps:
                amount = float(st_deps[0].amount)
                bank = st_deps[0].source_bank or m.bank_name or '-'
                account = st_deps[0].source_account_no or m.bank_account_number or '-'
            else:
                bal = balances_map.get(st.id, 0.0)
                amount = bal
                if bal > 0:
                    bank = m.bank_name or '-'
                    account = m.bank_account_number or '-'
                else:
                    bank = '-'
                    account = '-'
            applied[st.id] = amount
            banks[st.id] = bank
            accounts[st.id] = account
            
        member_applied[m.id] = applied
        member_banks[m.id] = banks
        member_accounts[m.id] = accounts

    pending_deposits = DepositRequest.query.filter_by(approval_status='PENDING').order_by(DepositRequest.created_at.desc()).all()
    pending_status_changes = MemberSavingBalance.query.filter(
        MemberSavingBalance.status.in_(['DEACTIVATION_PENDING', 'ACTIVATION_PENDING'])
    ).order_by(MemberSavingBalance.updated_at.desc()).all()

    return render_template('ledger.html',
                           current_user=current_user,
                           active_menu='savings',
                           page_title='Simpanan Anggota',
                           total_transactions=total_transactions,
                           total_transactions_bulan_ini=total_transactions_bulan_ini,
                           total_deposit_alltime=float(total_deposit_alltime),
                           total_balance=float(total_balance),
                           total_withdrawal=float(total_withdrawal),
                           pending_review=pending_review,
                           members=members,
                           saving_types=saving_types,
                           member_balances=member_balances,
                           member_applied=member_applied,
                           member_banks=member_banks,
                           member_accounts=member_accounts,
                           member_totals=member_totals,
                           member_statuses=member_statuses,
                           pending_deposits=pending_deposits,
                           pending_status_changes=pending_status_changes,
                           pagination=pagination,
                           search_query=search_query)

# ─────────────────────────────────────────────────────────────────────────────
# SAVINGS EXPORT HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _get_all_members_savings_data(search_query=None):
    """Return ALL members with their saving details (no pagination)."""
    query = Member.query.filter_by(status='AKTIF')
    if search_query:
        query = query.filter(
            (Member.full_name.ilike(f'%{search_query}%')) |
            (Member.member_no.ilike(f'%{search_query}%'))
        )
    all_members = query.order_by(Member.id.asc()).all()
    saving_types = SavingType.query.all()

    rows = []
    for m in all_members:
        deps = DepositRequest.query.filter(
            DepositRequest.member_id == m.id,
            DepositRequest.approval_status.in_(['APPROVED', 'PENDING'])
        ).order_by(DepositRequest.created_at.desc()).all()

        db_balances = MemberSavingBalance.query.filter_by(member_id=m.id).all()
        balances_map = {b.saving_type_id: float(b.balance) for b in db_balances}
        statuses_map = {b.saving_type_id: b.status or 'ACTIVE' for b in db_balances}
        total = sum(balances_map.values())

        row = {
            'member_no': m.member_no,
            'full_name': m.full_name,
            'total_saldo': total,
        }
        for st in saving_types:
            st_deps = [d for d in deps if d.saving_type_id == st.id]
            status = statuses_map.get(st.id, 'ACTIVE')
            
            # Map status
            status_label = 'Aktif'
            if status == 'INACTIVE':
                status_label = 'Nonaktif'
            elif status == 'DEACTIVATION_PENDING':
                status_label = 'Minta Nonaktif'
            elif status == 'ACTIVATION_PENDING':
                status_label = 'Minta Aktif'

            row[f'{st.name}_status'] = status_label
            if st_deps:
                row[f'{st.name}_nominal'] = float(st_deps[0].amount)
                row[f'{st.name}_bank'] = st_deps[0].source_bank or m.bank_name or '-'
                row[f'{st.name}_rekening'] = st_deps[0].source_account_no or m.bank_account_number or '-'
            else:
                bal = balances_map.get(st.id, 0.0)
                row[f'{st.name}_nominal'] = bal
                row[f'{st.name}_bank'] = (m.bank_name or '-') if bal > 0 else '-'
                row[f'{st.name}_rekening'] = (m.bank_account_number or '-') if bal > 0 else '-'
        rows.append(row)
    return rows, saving_types


@finance_bp.route('/finance/savings/excel')
def savings_export_excel():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    search_query = request.args.get('search', '').strip()
    rows, saving_types = _get_all_members_savings_data(search_query)

    # Build header columns
    headers = ['No. Anggota', 'Nama Lengkap']
    for st in saving_types:
        headers += [f'{st.name} - Nominal', f'{st.name} - Bank', f'{st.name} - No. Rekening', f'{st.name} - Status']
    headers.append('Total Saldo Aktif')

    # Build data rows
    data_rows = []
    for r in rows:
        row_data = [r['member_no'], r['full_name']]
        for st in saving_types:
            row_data += [
                r.get(f'{st.name}_nominal', 0),
                r.get(f'{st.name}_bank', '-'),
                r.get(f'{st.name}_rekening', '-'),
                r.get(f'{st.name}_status', 'Aktif'),
            ]
        row_data.append(r['total_saldo'])
        data_rows.append(row_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(data_rows, columns=headers)
        df.to_excel(writer, index=False, sheet_name='Simpanan Anggota')
        ws = writer.sheets['Simpanan Anggota']

        # Style header row
        header_fill = PatternFill('solid', fgColor='6e0b0b')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Style data rows with alternating colors
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill('solid', fgColor='FFF5F5') if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output.seek(0)
    filename = f'Simpanan_Anggota_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@finance_bp.route('/finance/savings/pdf')
def savings_export_pdf():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    search_query = request.args.get('search', '').strip()
    rows, saving_types = _get_all_members_savings_data(search_query)
    current_user = User.query.get(session['user_id'])
    return render_template('print_ledger.html',
                           rows=rows,
                           saving_types=saving_types,
                           current_user=current_user,
                           export_date=datetime.now().strftime('%d %B %Y, %H:%M'))


@finance_bp.route('/finance/deposit/<int:deposit_id>/approve', methods=['POST'])
def approve_deposit(deposit_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        deposit = DepositRequest.query.get(deposit_id)
        if not deposit or deposit.approval_status != 'PENDING':
            return jsonify({'success': False, 'error': 'Deposit tidak valid atau sudah diproses'}), 400
            
        deposit.approval_status = 'APPROVED'
        deposit.approved_by = session['user_id']
        deposit.approved_at = datetime.now(timezone.utc)

        # Ensure MemberSavingBalance record exists for this member and saving type
        balance_record = MemberSavingBalance.query.filter_by(
            member_id=deposit.member_id, saving_type_id=deposit.saving_type_id
        ).first()
        if not balance_record:
            balance_record = MemberSavingBalance(
                member_id=deposit.member_id,
                saving_type_id=deposit.saving_type_id,
                balance=0.0
            )
            db.session.add(balance_record)
            db.session.flush()
        
        # Notify member
        st = SavingType.query.get(deposit.saving_type_id)
        st_name = st.name if st else "Simpanan"
        Notification.create(
            member_id=deposit.member_id,
            title="Pengajuan Simpanan Disetujui",
            message=f"Pengajuan simpanan {st_name} Anda sebesar Rp {int(deposit.amount):,} telah disetujui oleh pengurus.".replace(',', '.'),
            notification_type="SAVING_APPROVED"
        )

        db.session.commit()
        return jsonify({'success': True, 'message': 'Pengajuan Buka Simpanan berhasil disetujui.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@finance_bp.route('/finance/deposit/<int:deposit_id>/reject', methods=['POST'])
def reject_deposit(deposit_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        deposit = DepositRequest.query.get(deposit_id)
        if not deposit or deposit.approval_status != 'PENDING':
            return jsonify({'success': False, 'error': 'Deposit tidak valid atau sudah diproses'}), 400
            
        data = request.get_json() or {}
        reason = data.get('reason', 'Ditolak oleh Admin')
        
        deposit.approval_status = 'REJECTED'
        deposit.rejection_reason = reason
        deposit.approved_by = session['user_id']
        deposit.approved_at = datetime.now(timezone.utc)
        
        # Notify member
        st = SavingType.query.get(deposit.saving_type_id)
        st_name = st.name if st else "Simpanan"
        Notification.create(
            member_id=deposit.member_id,
            title="Pengajuan Simpanan Ditolak",
            message=f"Pengajuan simpanan {st_name} Anda sebesar Rp {int(deposit.amount):,} ditolak. Alasan: {reason}".replace(',', '.'),
            notification_type="SAVING_REJECTED"
        )

        db.session.commit()
        return jsonify({'success': True, 'message': 'Setoran berhasil ditolak.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/finance/saving_balance/<int:member_id>/<int:st_id>/approve_status')
def approve_saving_status(member_id, st_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    action = request.args.get('action', 'deactivate')
    
    balance_record = MemberSavingBalance.query.filter_by(
        member_id=member_id, saving_type_id=st_id
    ).first()
    if not balance_record:
        flash('Saldo/Simpanan tidak ditemukan', 'error')
        return redirect(url_for('finance.savings'))
        
    st = SavingType.query.get(st_id)
    st_name = st.name if st else "Simpanan"
    
    if action == 'deactivate':
        balance_record.status = 'INACTIVE'
        msg = f"Permohonan penonaktifan Simpanan {st_name} Anda telah disetujui oleh pengurus."
        Notification.create(
            member_id=member_id,
            title="Penonaktifan Simpanan Disetujui",
            message=msg,
            notification_type="SAVING_APPROVED"
        )
        flash(f'Simpanan {st_name} berhasil dinonaktifkan.', 'success')
    else:
        balance_record.status = 'ACTIVE'
        msg = f"Permohonan pengaktifan kembali Simpanan {st_name} Anda telah disetujui oleh pengurus."
        Notification.create(
            member_id=member_id,
            title="Pengaktifan Simpanan Disetujui",
            message=msg,
            notification_type="SAVING_APPROVED"
        )
        flash(f'Simpanan {st_name} berhasil diaktifkan kembali.', 'success')
        
    db.session.commit()
    return redirect(url_for('finance.savings'))


@finance_bp.route('/finance/saving_balance/<int:member_id>/<int:st_id>/reject_status')
def reject_saving_status(member_id, st_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    balance_record = MemberSavingBalance.query.filter_by(
        member_id=member_id, saving_type_id=st_id
    ).first()
    if not balance_record:
        flash('Saldo/Simpanan tidak ditemukan', 'error')
        return redirect(url_for('finance.savings'))
        
    st = SavingType.query.get(st_id)
    st_name = st.name if st else "Simpanan"
    
    current_status = balance_record.status
    if current_status == 'DEACTIVATION_PENDING':
        balance_record.status = 'ACTIVE'
        msg = f"Permohonan penonaktifan Simpanan {st_name} Anda ditolak oleh pengurus."
        Notification.create(
            member_id=member_id,
            title="Penonaktifan Simpanan Ditolak",
            message=msg,
            notification_type="SAVING_REJECTED"
        )
        flash(f'Permohonan penonaktifan Simpanan {st_name} ditolak.', 'info')
    elif current_status == 'ACTIVATION_PENDING':
        balance_record.status = 'INACTIVE'
        msg = f"Permohonan pengaktifan kembali Simpanan {st_name} Anda ditolak oleh pengurus."
        Notification.create(
            member_id=member_id,
            title="Pengaktifan Simpanan Ditolak",
            message=msg,
            notification_type="SAVING_REJECTED"
        )
        flash(f'Permohonan pengaktifan kembali Simpanan {st_name} ditolak.', 'info')
        
    db.session.commit()
    return redirect(url_for('finance.savings'))


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PAYROLL UPLOAD & TEMPLATE
# ─────────────────────────────────────────────────────────────────────────────
import pandas as pd
import io
from flask import send_file

@finance_bp.route('/finance/payroll/download_template')
def download_payroll_template():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    active_members = Member.query.filter_by(status='AKTIF').all()
    data = []
    for m in active_members:
        data.append({
            'member_no': m.member_no,
            'member_name': m.full_name,
            'saving_type_code': 'SW', # Default Wajib
            'amount': 500000          # Default 500k
        })
        
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Payroll_Template')
        
    output.seek(0)
    filename = f"Template_Payroll_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@finance_bp.route('/finance/payroll/upload_process', methods=['POST'])
def upload_process_payroll():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Tidak ada file yang diunggah'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'File kosong'}), 400
        
    try:
        df = pd.read_excel(file)
        
        # Cari kolom nominal dinamis yang berakhiran ' - Nominal'
        nominal_cols = [c for c in df.columns if c.endswith(' - Nominal')]
        if not nominal_cols:
            return jsonify({'success': False, 'error': 'Tidak ada kolom nominal simpanan (misal: "Simpanan Pokok - Nominal") di Excel.'}), 400
            
        # Peta nama simpanan -> objek SavingType
        saving_types_by_name = {st.name: st for st in SavingType.query.all()}
        col_to_saving_type = {}
        for col in nominal_cols:
            st_name = col.split(' - Nominal')[0].strip()
            if st_name in saving_types_by_name:
                col_to_saving_type[col] = saving_types_by_name[st_name]

        if not col_to_saving_type:
            return jsonify({'success': False, 'error': 'Tidak ada nama jenis simpanan dalam kolom Excel yang cocok dengan database.'}), 400

        batch_code = "PRX-" + datetime.now().strftime("%Y%m") + "-" + secrets.token_hex(2).upper()
        
        new_batch = PayrollBatch(
            batch_code=batch_code,
            period_month=datetime.now().month,
            period_year=datetime.now().year,
            total_amount=0,       # Akan diupdate setelah selesai loop
            total_members=0,      # Akan diupdate setelah selesai loop
            distribution_status='PROCESSING',
            validation_status='SUCCESS',
            uploaded_by=session['user_id'],
            uploaded_at=datetime.now(timezone.utc),
        )
        db.session.add(new_batch)
        db.session.flush()
        
        success_count = 0
        failed_count = 0
        total_amount = 0
        
        for index, row in df.iterrows():
            member = None
            try:
                # Cari kunci anggota di Excel: 'No. Anggota' atau 'member_no'
                member_no = row.get('No. Anggota')
                if pd.isna(member_no):
                    member_no = row.get('member_no')
                
                if pd.isna(member_no):
                    continue
                    
                member_no = str(member_no).strip()
                member = Member.query.filter_by(member_no=member_no).first()
                if not member:
                    raise Exception(f"Anggota {member_no} tidak ditemukan.")
                
                member_has_transaction = False
                for col, st in col_to_saving_type.items():
                    val = row.get(col, 0)
                    if pd.isna(val):
                        val = 0
                    amount = float(val)
                    if amount <= 0:
                        continue
                    
                    balance_record = MemberSavingBalance.query.filter_by(
                        member_id=member.id, saving_type_id=st.id
                    ).first()
                    if balance_record and balance_record.status == 'INACTIVE':
                        continue
                    if not balance_record:
                        balance_record = MemberSavingBalance(
                            member_id=member.id, saving_type_id=st.id, balance=0
                        )
                        db.session.add(balance_record)
                        db.session.flush()

                    balance_before = float(balance_record.balance)
                    balance_after = balance_before + amount

                    balance_record.balance = balance_after
                    balance_record.last_transaction_at = datetime.now(timezone.utc)

                    detail = PayrollBatchDetail(
                        payroll_batch_id=new_batch.id,
                        member_id=member.id,
                        saving_type_id=st.id,
                        amount=amount,
                        distribution_status='SUCCESS'
                    )
                    db.session.add(detail)
                    db.session.flush()

                    trx = SavingTransaction(
                        member_id=member.id,
                        payroll_batch_detail_id=detail.id,
                        saving_type_id=st.id,
                        transaction_type='DEBIT',
                        amount=amount,
                        balance_before=balance_before,
                        balance_after=balance_after,
                        transaction_source='PAYROLL',
                        reference_number="TRX-PAYX-" + secrets.token_hex(6).upper(),
                        transaction_date=datetime.now(timezone.utc),
                        transaction_status='SUCCESS',
                        processed_by=session['user_id'],
                        processed_at=datetime.now(timezone.utc),
                        description=f'Pencatatan {st.name} oleh pengurus koperasi'
                    )
                    db.session.add(trx)
                    
                    # Kirim notifikasi
                    st_display_name = st.name
                    if st_display_name.lower().startswith('simpanan '):
                        st_display_name = st_display_name[9:].strip()
                    Notification.create(
                        member_id=member.id,
                        title="Simpanan Masuk (Payroll)",
                        message=f"Simpanan {st_display_name} sebesar Rp {int(amount):,} telah berhasil diproses oleh pengurus koperasi dan ditambahkan ke saldo simpanan Anda.".replace(',', '.'),
                        notification_type="PAYROLL"
                    )

                    total_amount += amount
                    member_has_transaction = True
                
                if member_has_transaction:
                    success_count += 1
                else:
                    # Anggota ada tapi tidak ada nominal simpanan > 0
                    pass
                    
            except Exception as row_err:
                failed_count += 1
                print(f"[EXCEL PAYROLL] Baris {int(float(index))+2} gagal: {row_err}")
                
                if member:
                    detail = PayrollBatchDetail(
                        payroll_batch_id=new_batch.id,
                        member_id=member.id,
                        saving_type_id=None,
                        amount=0,
                        distribution_status='FAILED'
                    )
                    db.session.add(detail)
        
        new_batch.total_amount = total_amount
        new_batch.total_members = success_count
        new_batch.distribution_status = 'SUCCESS' if failed_count == 0 else 'PARTIAL'
        new_batch.success_count = success_count
        new_batch.failed_count = failed_count
        new_batch.processed_at = datetime.now(timezone.utc)

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Payroll Excel berhasil diproses. Sukses: {success_count} anggota, Gagal: {failed_count}.',
            'batch_code': batch_code,
            'success_count': success_count,
            'failed_count': failed_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ─────────────────────────────────────────────────────────────────────────────
# PENARIKAN
# ─────────────────────────────────────────────────────────────────────────────
@finance_bp.route('/finance/withdrawals')
def withdrawals():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    current_user = User.query.get(session['user_id'])

    reqs = WithdrawalRequest.query.order_by(WithdrawalRequest.request_date.desc()).all()
    stats = {
        'pending': WithdrawalRequest.query.filter_by(approval_status='PENDING').count(),
        'approved_today': WithdrawalRequest.query.filter_by(approval_status='APPROVED').count(),
        'completed': WithdrawalRequest.query.filter_by(approval_status='COMPLETED').count(),
        'rejected': WithdrawalRequest.query.filter_by(approval_status='REJECTED').count()
    }
    return render_template('withdrawal.html', current_user=current_user, withdrawals=reqs, stats=stats,
                           active_menu='withdrawals', page_title='Penarikan Simpanan')


@finance_bp.route('/finance/withdraw/<int:withdraw_id>/accept', methods=['POST'])
def accept_withdrawal(withdraw_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        withdraw = WithdrawalRequest.query.get(withdraw_id)
        if not withdraw or withdraw.approval_status != 'PENDING':
            return jsonify({'success': False, 'error': 'Penarikan tidak valid atau sudah diproses'}), 400
            
        withdraw.approval_status = 'ACCEPTED'
        
        st_id = 2 
        if withdraw.processing_notes and "Saving Type ID:" in withdraw.processing_notes:
            try:
                st_id = int(withdraw.processing_notes.split("Saving Type ID:")[1].split(".")[0].strip())
            except:
                pass
        st = SavingType.query.get(st_id)
        st_name = st.name if st else "Simpanan"

        Notification.create(
            member_id=withdraw.member_id,
            title="Penarikan Diterima",
            message="Pengajuan penarikan Anda telah diterima oleh pengurus dan akan segera diproses.",
            notification_type="WITHDRAWAL_ACCEPTED"
        )

        db.session.commit()
        return jsonify({'success': True, 'message': 'Pengajuan Penarikan berhasil diterima.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/finance/withdraw/<int:withdraw_id>/process', methods=['POST'])
def process_withdrawal(withdraw_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        withdraw = WithdrawalRequest.query.get(withdraw_id)
        if not withdraw or withdraw.approval_status != 'ACCEPTED':
            return jsonify({'success': False, 'error': 'Penarikan belum diterima atau sudah diproses'}), 400
            
        withdraw.approval_status = 'PROCESSING'
        withdraw.approved_by = session['user_id']
        withdraw.approved_at = datetime.now(timezone.utc)

        Notification.create(
            member_id=withdraw.member_id,
            title="Penarikan Sedang Diproses",
            message="Pengajuan penarikan Anda sedang diproses oleh pengurus.",
            notification_type="WITHDRAWAL_PROCESSING"
        )

        db.session.commit()
        return jsonify({'success': True, 'message': 'Proses transfer penarikan telah dimulai.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/finance/withdraw/<int:withdraw_id>/complete', methods=['POST'])
def complete_withdrawal(withdraw_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        withdraw = WithdrawalRequest.query.get(withdraw_id)
        if not withdraw or withdraw.approval_status != 'PROCESSING':
            return jsonify({'success': False, 'error': 'Penarikan tidak valid atau belum dalam status PROCESSING'}), 400
            
        file = request.files.get('transfer_proof')
        if not file or file.filename == '':
            return jsonify({'success': False, 'error': 'Bukti transfer wajib diunggah sebelum menyelesaikan proses'}), 400

        custom_ref = request.form.get('reference_number', '').strip()

        import os
        from werkzeug.utils import secure_filename
        upload_folder = os.path.join('static', 'uploads', 'proofs')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder, exist_ok=True)
            
        filename = secure_filename(f"proof_wd_{withdraw_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
        filepath = os.path.join(upload_folder, filename)
        file.save(filepath)
        proof_url = f"/static/uploads/proofs/{filename}"

        st_id = 2 
        if withdraw.processing_notes and "Saving Type ID:" in withdraw.processing_notes:
            try:
                st_id = int(withdraw.processing_notes.split("Saving Type ID:")[1].split(".")[0].strip())
            except:
                pass

        st = SavingType.query.get(st_id)
        st_name = st.name if st else "Simpanan"
        
        balance_record = MemberSavingBalance.query.filter_by(member_id=withdraw.member_id, saving_type_id=st_id).first()
        if not balance_record or float(balance_record.balance) < float(withdraw.amount):
            return jsonify({'success': False, 'error': 'Saldo tidak mencukupi'}), 400

        balance_before = float(balance_record.balance)
        balance_after = balance_before - float(withdraw.amount)
        
        withdraw.approval_status = 'APPROVED'
        withdraw.completed_at = datetime.now(timezone.utc)
        withdraw.transfer_proof = proof_url

        balance_record.balance = balance_after
        balance_record.last_transaction_at = datetime.now(timezone.utc)

        # Gunakan nomor referensi kustom atau generate otomatis
        if custom_ref:
            # pastikan unik
            existing = SavingTransaction.query.filter_by(reference_number=custom_ref).first()
            if existing:
                return jsonify({'success': False, 'error': f'Nomor referensi transaksi "{custom_ref}" sudah digunakan'}), 400
            ref_num = custom_ref
        else:
            import secrets
            ref_num = "TRX-WD-" + secrets.token_hex(6).upper()

        trx = SavingTransaction(
            member_id=withdraw.member_id,
            saving_type_id=st_id,
            transaction_type='CREDIT', 
            amount=withdraw.amount,
            balance_before=balance_before,
            balance_after=balance_after,
            transaction_source='WITHDRAWAL',
            reference_number=ref_num,
            transaction_date=datetime.now(timezone.utc),
            transaction_status='SUCCESS',
            processed_by=session['user_id'],
            processed_at=datetime.now(timezone.utc),
            description=f'Pencatatan penarikan {st_name} oleh pengurus koperasi'
        )
        db.session.add(trx)
        db.session.flush()

        withdraw.saving_transaction_id = trx.id

        Notification.create(
            member_id=withdraw.member_id,
            title="Penarikan Berhasil",
            message=f"Pengajuan penarikan Anda telah berhasil diselesaikan. Saldo Anda telah terpotong Rp {int(withdraw.amount):,}.".replace(',', '.'),
            notification_type="WITHDRAWAL_APPROVED"
        )

        db.session.commit()
        return jsonify({'success': True, 'message': 'Proses Penarikan berhasil diselesaikan.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@finance_bp.route('/finance/withdraw/<int:withdraw_id>/reject', methods=['POST'])
def reject_withdrawal(withdraw_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        withdraw = WithdrawalRequest.query.get(withdraw_id)
        if not withdraw or withdraw.approval_status not in ['PENDING', 'PROCESSING']:
            return jsonify({'success': False, 'error': 'Penarikan tidak valid atau sudah diproses'}), 400
            
        data = request.get_json() or {}
        reason = data.get('reason', 'Ditolak oleh Admin')

        st_id = 2
        if withdraw.processing_notes and "Saving Type ID:" in withdraw.processing_notes:
            try:
                st_id = int(withdraw.processing_notes.split("Saving Type ID:")[1].split(".")[0].strip())
            except:
                pass

        st = SavingType.query.get(st_id)
        st_name = st.name if st else "Simpanan"
        
        withdraw.approval_status = 'REJECTED'
        withdraw.rejection_reason = reason
        withdraw.approved_by = session['user_id']
        withdraw.approved_at = datetime.now(timezone.utc)

        Notification.create(
            member_id=withdraw.member_id,
            title="Penarikan Ditolak",
            message=f"Pengajuan penarikan {st_name} Anda sebesar Rp {int(withdraw.amount):,} ditolak. Alasan: {reason}".replace(',', '.'),
            notification_type="WITHDRAWAL_REJECTED"
        )

        db.session.commit()
        return jsonify({'success': True, 'message': 'Pengajuan Penarikan berhasil ditolak.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
